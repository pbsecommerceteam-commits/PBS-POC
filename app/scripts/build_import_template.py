"""
Builds the Shelfline data-import template (app/public/templates/
shelfline-data-import-template.xlsx): a 4-tab .xlsx (Content / Price /
Share Of Search / MAP Price) matching the exact real crawl format
build_mock_data.py reads, pre-filled with the current real 117-SKU
September 2022 crawl -- read straight from the actual source workbooks on
disk, not reconstructed from mockData.ts's lossier weekly-bucketed data, so
the prefilled rows are genuinely the same data at full original fidelity.

On top of the raw source format, this adds:
  - A "Company" column on every tab (first column), for a future upload to
    carry another client's data through this same template.
  - Renames: SPB Url/Spb url -> Url, Category name -> Category/account
    name, Vendor stock no -> SKU (see build_mock_data.py's get_any() for
    why the ETL script still accepts the old names too).
  - Drops columns the ETL script now computes itself instead of reading
    (Title/Description "no of chars", Bullet N length) and Ingredients
    list (removed from the app entirely).
  - Adds "Other Seller 1..10 Name/Price" (Price tab, blank on this crawl --
    never captured) and "Brand_1..65" per Share Of Search rank (filled in
    for the existing rows via a best-effort name match against this
    catalog's own known brands, falling back to the result's first word --
    approximate by nature, see the Instructions tab).
  - Backfills a Url + SKU column on MAP Price (joined from the Price tab by
    Retailer site + Retailer id) so a MAP row is identifiable without
    cross-referencing another tab.
  - Reorders every tab so the real identifying columns (Company, SKU,
    Retailer site, Retailer id, Url) sit leftmost, then Crawl date, then
    everything else -- with the identifier group and the date column each
    given their own header color so the two are visually distinct from the
    rest.

Usage: python build_import_template.py [source-dir] [output-path]
Defaults match this machine's real source-file locations.
"""
import re
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC_DIR = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\YASH\Desktop\Data PBS_POC"
MAIN_WB = SRC_DIR + r"\Main Working File.xlsx"
MAP_WB = SRC_DIR + r"\Map price.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\YASH\Desktop\PBS-POC\app\public\templates\shelfline-data-import-template.xlsx"

COMPANY_DEFAULT = "Perfality"
OTHER_SELLER_SLOTS = 10
SOS_URL_SLOTS = 65

IDENTIFIER_FILL = PatternFill("solid", fgColor="000000")
IDENTIFIER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATE_FILL = PatternFill("solid", fgColor="C55A11")  # orange
DATE_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # existing dark navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
COMPANY_FILL = PatternFill("solid", fgColor="FFF6D8")
COMPANY_FONT = Font(bold=True, size=10)

RENAME = {"SPB Url": "Url", "Spb url": "Url", "Category name": "Category/account name", "Vendor stock no": "SKU"}
DROP = {"Title no of chars", "Description no of chars", "Ingredients list"} | {f"Bullet {i} length" for i in range(1, 11)}


def load_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return header, [dict(zip(header, r)) for r in rows[1:] if r and any(c is not None for c in r)]


def renamed(name):
    return RENAME.get(name, name)


def reorder(headers, identifier_cols, date_col):
    """Puts the real identifier columns first, then the date column, then
    everything else in its original relative order -- so every tab reads
    "what is this row" before "when was it observed" before the rest."""
    front = [c for c in identifier_cols if c in headers]
    if date_col and date_col in headers:
        front.append(date_col)
    rest = [h for h in headers if h not in front]
    return front + rest


def other_seller_cols():
    cols = []
    for i in range(1, OTHER_SELLER_SLOTS + 1):
        cols += [f"Other Seller {i} Name", f"Other Seller {i} Price"]
    return cols


def sos_slot_cols():
    cols = []
    for i in range(1, SOS_URL_SLOTS + 1):
        cols += [f"Url_{i}", f"Url_{i}_Sponsored", f"Product_name_{i}", f"Brand_{i}"]
    return cols


# ── best-effort Brand_N inference for the existing real SOS rows ───────────
def build_brand_matcher(content_rows):
    known = sorted({r["Brand"] for r in content_rows if r.get("Brand")}, key=len, reverse=True)
    lowered = [(b.lower(), b) for b in known]

    def guess(name):
        if not name:
            return None
        n = str(name)
        nl = n.lower()
        for bl, canon in lowered:
            if nl.startswith(bl):
                return canon
        # Fallback: no known brand matched (almost every competitor result)
        # -- best guess is the product's own first word/words, since that's
        # how most listings are titled ("Purina ONE SmartBlend...",
        # "Hartz UltraGuard..."). Approximate by construction; see the
        # Instructions tab.
        words = re.findall(r"[A-Za-z0-9&'.+-]+", n)
        if not words:
            return None
        guess_words = [words[0]]
        if len(words) > 1 and (words[1][:1].isupper() or words[1].isupper()):
            guess_words.append(words[1])
        return " ".join(guess_words)

    return guess


def copy_sheet(dst_wb, sheet_name, headers, rows, identifier_cols, date_col, company_default=COMPANY_DEFAULT):
    """headers: final ordered header list (post-rename, minus dropped cols,
    minus Company -- Company is always prepended). rows: list of dicts
    already keyed by FINAL header names (renamed/added columns already
    computed by the caller)."""
    dst_ws = dst_wb.create_sheet(sheet_name)
    full_headers = ["Company"] + headers
    dst_ws.append(full_headers)
    for r in rows:
        dst_ws.append([r.get("Company", company_default)] + [r.get(h) for h in headers])

    ident_set = {"Company"} | set(identifier_cols)
    for ci, h in enumerate(full_headers, start=1):
        c = dst_ws.cell(row=1, column=ci)
        if h in ident_set:
            c.fill = IDENTIFIER_FILL
            c.font = IDENTIFIER_FONT
        elif date_col and h == date_col:
            c.fill = DATE_FILL
            c.font = DATE_FONT
        else:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")

    for ri in range(2, dst_ws.max_row + 1):
        c = dst_ws.cell(row=ri, column=1)
        c.fill = COMPANY_FILL
        c.font = COMPANY_FONT

    dst_ws.freeze_panes = get_column_letter(len(ident_set) + (2 if date_col else 1)) + "2"
    dst_ws.column_dimensions["A"].width = 16
    for ci in range(2, min(len(full_headers), 40) + 1):
        dst_ws.column_dimensions[get_column_letter(ci)].width = 15
    return dst_ws, len(rows)


def main():
    src_main = openpyxl.load_workbook(MAIN_WB, read_only=True, data_only=True)
    src_map = openpyxl.load_workbook(MAP_WB, read_only=True, data_only=True)

    content_header, content_rows = load_rows(src_main["Content"])
    price_header, price_rows = load_rows(src_main["Price"])
    sos_header, sos_rows = load_rows(src_main["Share Of Search"])
    _, map_rows = load_rows(src_map["Sheet3"])  # Retailer site, Retailer id, Map Price

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    counts = {}

    # ── Content ──────────────────────────────────────────────────────────
    content_headers = reorder([renamed(h) for h in content_header if h not in DROP],
                               identifier_cols=["SKU", "Retailer site", "Retailer id", "Url"], date_col="Crawl date")
    content_out_rows = [{renamed(k): v for k, v in r.items() if k not in DROP} for r in content_rows]
    for r in content_out_rows:
        r["Company"] = COMPANY_DEFAULT
    _, n = copy_sheet(out_wb, "Content", content_headers, content_out_rows,
                       identifier_cols=["SKU", "Retailer site", "Retailer id", "Url"], date_col="Crawl date")
    counts["Content"] = n

    # ── Price (+ Other Seller 1..10 Name/Price, blank -- never captured) ──
    price_headers = reorder([renamed(h) for h in price_header] + other_seller_cols(),
                             identifier_cols=["SKU", "Retailer site", "Retailer id", "Url"], date_col="Crawl date")
    price_out_rows = []
    for r in price_rows:
        row = {renamed(k): v for k, v in r.items()}
        row["Company"] = COMPANY_DEFAULT
        price_out_rows.append(row)
    _, n = copy_sheet(out_wb, "Price", price_headers, price_out_rows,
                       identifier_cols=["SKU", "Retailer site", "Retailer id", "Url"], date_col="Crawl date")
    counts["Price"] = n

    # ── Share Of Search (+ Brand_1..65, best-effort inferred) ────────────
    guess_brand = build_brand_matcher(content_rows)
    sos_slots = sos_slot_cols()
    sos_headers = ["site", "Crawl_date", "keyword"] + sos_slots
    sos_out_rows = []
    for r in sos_rows:
        row = {"site": r.get("site"), "keyword": r.get("keyword"), "Crawl_date": r.get("Crawl_date"), "Company": COMPANY_DEFAULT}
        for i in range(1, SOS_URL_SLOTS + 1):
            row[f"Url_{i}"] = r.get(f"Url_{i}")
            row[f"Url_{i}_Sponsored"] = r.get(f"Url_{i}_Sponsored")
            pname = r.get(f"Product_name_{i}")
            row[f"Product_name_{i}"] = pname
            row[f"Brand_{i}"] = guess_brand(pname) if pname else None
        sos_out_rows.append(row)
    _, n = copy_sheet(out_wb, "Share Of Search", sos_headers, sos_out_rows,
                       identifier_cols=["site"], date_col="Crawl_date")
    counts["Share Of Search"] = n

    # ── MAP Price (+ Url, SKU backfilled by joining the Price tab) ───────
    by_site_id = {}
    for r in price_rows:
        key = (r.get("Retailer site"), str(r.get("Retailer id")))
        by_site_id[key] = r  # last (most recent) row per key wins, same as build_mock_data.py's latest_price_row
    map_out_rows = []
    for r in map_rows:
        key = (r.get("Retailer site"), str(r.get("Retailer id")))
        joined = by_site_id.get(key, {})
        map_out_rows.append({
            "Company": COMPANY_DEFAULT,
            "SKU": joined.get("Vendor stock no"),
            "Retailer site": r.get("Retailer site"),
            "Retailer id": r.get("Retailer id"),
            "Url": joined.get("Spb url"),
            "Map Price": r.get("Map Price"),
        })
    _, n = copy_sheet(out_wb, "MAP Price", ["SKU", "Retailer site", "Retailer id", "Url", "Map Price"], map_out_rows,
                       identifier_cols=["SKU", "Retailer site", "Retailer id", "Url"], date_col=None)
    counts["MAP Price"] = n

    # ── Instructions tab (first) ────────────────────────────────────────────
    instr = out_wb.create_sheet("Instructions", 0)
    instr.sheet_view.showGridLines = False
    instr.column_dimensions["A"].width = 112
    title_font = Font(bold=True, size=14)
    h_font = Font(bold=True, size=11)
    body_font = Font(size=10.5)
    wrap = Alignment(wrap_text=True, vertical="top")

    lines = [
        ("Shelfline data import template", title_font),
        ("", body_font),
        ("Pre-filled with the real September 2022 crawl currently powering the dashboard, so you can see", body_font),
        ("exactly the format each column expects. Fill in new or changed rows using the same columns, then", body_font),
        ("upload this file from Workspace > Data Import in the app.", body_font),
        ("", body_font),
        ("Column layout, every tab", h_font),
        ("  The real identifying columns (Company, SKU, Retailer site, Retailer id, Url) are always leftmost,", body_font),
        ("  shaded black. Crawl date (or Crawl_date) comes right after, shaded orange. Everything else follows", body_font),
        ("  in its original order. Do not rename, reorder, or delete a header -- the importer matches columns", body_font),
        ("  by name.", body_font),
        ("", body_font),
        ("4 data tabs", h_font),
        ("  - Content: one row per SKU per crawl date -- title, description, bullets, images, videos,", body_font),
        ("    variations, rating, reviews. Title/description/bullet character counts are no longer entered", body_font),
        ("    here -- the pipeline computes them from the text itself.", body_font),
        ("  - Price: one row per SKU per crawl date -- list/current/subscription price, stock status, buy box", body_font),
        ("    seller, coupon, and up to 10 Other Seller Name/Price pairs for competing (non-buy-box) offers.", body_font),
        ("  - Share Of Search: one row per (keyword, retailer, crawl date) with up to 65 ranked results, each", body_font),
        ("    with its Url, whether it's sponsored, the product name, and a Brand. Optional -- leave blank if", body_font),
        ("    you don't track this. Brand on the prefilled rows is a best-effort guess (matched against this", body_font),
        ("    catalog's own brands, or the result's first word/words otherwise) -- correct it where you know", body_font),
        ("    the real brand; it was never re-crawled.", body_font),
        ("  - MAP Price: one row per SKU with its Minimum Advertised Price policy value, plus SKU/Url for", body_font),
        ("    reference. Optional.", body_font),
        ("", body_font),
        ("SKU column", h_font),
        ("  The vendor's own stock-keeping unit (was \"Vendor stock no\"). When the same SKU appears under 2+", body_font),
        ("  different retailers, the app now treats that as the same real product sold in more than one place", body_font),
        ("  (a more reliable signal than matching by name) and surfaces it on that product's page.", body_font),
        ("", body_font),
        ("Company column", h_font),
        ("  Every tab's first column. Existing rows are filled in with \"" + COMPANY_DEFAULT + "\", the current client.", body_font),
        ("  To load another client's real data using this same template and pipeline, give their rows their own", body_font),
        ("  Company name -- everything downstream keys off (Company, Retailer site, Retailer id / SKU), so", body_font),
        ("  different clients' SKUs never collide.", body_font),
        ("", body_font),
        ("Update vs. Add new data", h_font),
        ("  - Update existing data: re-crawled rows for SKUs already in the dashboard (same Company + Retailer", body_font),
        ("    site + Retailer id). Refreshes their price/content/rating history.", body_font),
        ("  - Add new data: brand-new SKUs, a new retailer, or a new Company entirely. Appended rather than", body_font),
        ("    replacing. Pick the mode in the app when you upload -- it does not change anything in this file.", body_font),
        ("", body_font),
        ("Formats", h_font),
        ("  - Crawl date / Crawl_date: a real date (YYYY-MM-DD, e.g. 2022-09-08).", body_font),
        ("  - Retailer site / site: one of amazon.com, chewy.com, walmart.com, homedepot.com, petsmart.com,", body_font),
        ("    lowes.com, petco.com.", body_font),
        ("  - Retailer id: the retailer's own native product id (ASIN, item id, etc.) -- must match exactly", body_font),
        ("    across tabs for the same SKU.", body_font),
        ("  - Prices: plain numbers, no currency symbol (12.99, not $12.99).", body_font),
        ("  - Rating: 0-5. Enhanced content: Yes or No.", body_font),
        ("", body_font),
        ("The app checks every row on upload and shows any errors (missing Company, unknown Retailer site,", body_font),
        ("non-numeric price, etc.) before anything is applied, so mistakes get caught at upload time -- fix the", body_font),
        ("flagged cells in this file and re-upload.", body_font),
        ("", body_font),
        ("Row counts in this prefilled copy: Content " + str(counts["Content"]) + ", Price " + str(counts["Price"]) +
         ", Share Of Search " + str(counts["Share Of Search"]) + ", MAP Price " + str(counts["MAP Price"]) + ".", body_font),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = instr.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = wrap

    out_wb.save(OUT)
    print("Wrote", OUT)
    print(counts)


if __name__ == "__main__":
    main()
